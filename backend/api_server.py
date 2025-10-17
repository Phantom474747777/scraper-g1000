"""
Scraper G1000 - Flask REST API Server
Provides HTTP endpoints for frontend and serves static UI files
"""

import sys
import os
import asyncio
import threading
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.profile_manager import ProfileManager
from src.database import LeadsDatabase
from src.scraper_free_bypass import scrape_yellowpages_free
from src.zip_lookup import get_zips_in_radius

# Determine UI directory
BASE_DIR = Path(__file__).parent.parent
UI_DIR = BASE_DIR / "scraper-g1000-tauri" / "src"
if not UI_DIR.exists():
    UI_DIR = BASE_DIR / "app"

# Create Flask app with static folder configured
app = Flask(__name__, static_folder=str(UI_DIR), static_url_path='')
CORS(app)  # Enable CORS for frontend

print(f"[Backend] Serving UI from: {UI_DIR}")
print(f"[Backend] UI exists: {UI_DIR.exists()}")

# Global state
profile_manager = ProfileManager()
scraping_state = {
    'active': False,
    'paused': False,
    'current_zip': None,
    'current_category': None,
    'total_leads': 0,
    'progress': 0,
    'stop_requested': False,
    'logs': [],  # Real-time log buffer
    'current_page': 0,
    'max_pages': 2
}

# === STATIC FILE SERVING ===

@app.route('/')
def serve_root():
    """Serve index.html at root"""
    return app.send_static_file('index.html')

@app.route('/js/<path:filename>')
def serve_js(filename):
    """Serve JavaScript files"""
    return send_from_directory(str(UI_DIR / 'js'), filename)

@app.route('/styles/<path:filename>')
def serve_styles(filename):
    """Serve CSS files"""
    return send_from_directory(str(UI_DIR / 'styles'), filename)

# === HEALTH CHECK ===

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'ok', 'message': 'Scraper G1000 Backend Running'}), 200

# === PROFILE MANAGEMENT ===

@app.route('/api/profiles', methods=['GET'])
def get_profiles():
    """Get all profiles with real lead counts from database"""
    try:
        profiles = profile_manager.get_all_profiles()
        print(f"[DEBUG] Loaded {len(profiles)} profiles", flush=True)

        result_profiles = []
        for p in profiles:
            print(f"[DEBUG] Processing profile: {p.name}")

            try:
                # Get database path
                db_path = p.get_database_path()
                print(f"[DEBUG] Database path: {db_path}")

                # Test if file exists
                import os
                if not os.path.exists(db_path):
                    print(f"[ERROR] Database file not found: {db_path}")
                    raise FileNotFoundError(f"Database not found: {db_path}")

                # Count valid leads
                import sqlite3
                print(f"[DEBUG] Connecting to database...")
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()

                # Count valid
                print(f"[DEBUG] Counting valid leads...")
                cursor.execute('SELECT COUNT(*) FROM leads WHERE phone NOT IN ("N/A", "") AND phone IS NOT NULL AND name NOT LIKE "%![%" AND name NOT LIKE "%[Website%" AND name NOT LIKE "%About Search%"')
                valid_count = cursor.fetchone()[0]
                print(f"[DEBUG] Valid leads: {valid_count}")

                # Count total
                print(f"[DEBUG] Counting total leads...")
                cursor.execute('SELECT COUNT(*) FROM leads')
                total_count = cursor.fetchone()[0]
                print(f"[DEBUG] Total leads: {total_count}")

                conn.close()

                result_profiles.append({
                    'id': p.profile_id,
                    'name': p.name,
                    'icon': p.icon,
                    'businessType': p.business_type,
                    'totalLeads': valid_count,
                    'flaggedLeads': total_count - valid_count,
                    'lastUsed': p.created_at
                })
                print(f"[DEBUG] Added profile {p.name} with {valid_count} valid leads, {total_count - valid_count} flagged")

            except Exception as db_error:
                print(f"[ERROR] Database error for profile {p.name}: {db_error}")
                import traceback
                traceback.print_exc()
                # Add profile with 0 leads on error
                result_profiles.append({
                    'id': p.profile_id,
                    'name': p.name,
                    'icon': p.icon,
                    'businessType': p.business_type,
                    'totalLeads': 0,
                    'flaggedLeads': 0,
                    'error': str(db_error),
                    'lastUsed': p.created_at
                })

        print(f"[DEBUG] Returning {len(result_profiles)} profiles")
        return jsonify({
            'success': True,
            'profiles': result_profiles
        }), 200

    except Exception as e:
        print(f"[ERROR] Get profiles outer exception: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/profiles', methods=['POST'])
def create_profile():
    """Create a new profile"""
    try:
        data = request.json
        profile = profile_manager.create_profile(
            name=data['name'],
            icon=data.get('icon', '📊'),
            business_type=data.get('businessType', ''),
            default_city=data.get('defaultCity', ''),
            default_state=data.get('defaultState', ''),
            categories=data.get('categories', [])
        )
        return jsonify({
            'success': True,
            'profileId': profile.profile_id
        }), 201
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/zip-lookup', methods=['POST'])
def zip_lookup():
    """Find ZIP codes within radius of a city"""
    try:
        data = request.json
        city = data.get('city')
        state = data.get('state')
        radius = int(data.get('radius', 50))

        if not city or not state:
            return jsonify({'success': False, 'error': 'City and state are required'}), 400

        # Suppress print output to avoid Windows encoding issues with emojis
        import io
        import contextlib

        # Redirect stdout to suppress emoji print statements
        f = io.StringIO()
        with contextlib.redirect_stdout(f):
            # Use existing zip lookup function
            zips = get_zips_in_radius(city, state, radius)

        print(f"[ZIP Lookup] Found {len(zips)} ZIPs near {city}, {state}")

        return jsonify({
            'success': True,
            'zips': zips,
            'count': len(zips)
        }), 200

    except Exception as e:
        print(f"[ZIP Lookup] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# === SCRAPING OPERATIONS ===

@app.route('/api/scrape/start', methods=['POST'])
def start_scraping():
    """Start a scraping job"""
    global scraping_state

    if scraping_state['active']:
        return jsonify({'success': False, 'error': 'Scraping already active'}), 400

    try:
        data = request.json
        profile_id = data.get('profileId')
        zip_code = data.get('zipCode')
        category = data.get('category')
        max_pages = data.get('maxPages', 2)

        # Reset state
        scraping_state.update({
            'active': True,
            'paused': False,
            'current_zip': zip_code,
            'current_category': category,
            'total_leads': 0,
            'progress': 0,
            'stop_requested': False,
            'logs': [],
            'current_page': 0,
            'max_pages': max_pages
        })

        # Start scraping in background thread
        thread = threading.Thread(
            target=run_scrape_job,
            args=(profile_id, zip_code, category, max_pages),
            daemon=True
        )
        thread.start()

        return jsonify({
            'success': True,
            'message': 'Scraping started',
            'status': scraping_state
        }), 200

    except Exception as e:
        scraping_state['active'] = False
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/scrape/pause', methods=['POST'])
def pause_scraping():
    """Pause scraping"""
    global scraping_state

    if not scraping_state['active']:
        return jsonify({'success': False, 'error': 'No active scrape'}), 400

    scraping_state['paused'] = True
    return jsonify({'success': True, 'message': 'Scraping paused'}), 200

@app.route('/api/scrape/resume', methods=['POST'])
def resume_scraping():
    """Resume scraping"""
    global scraping_state

    if not scraping_state['active']:
        return jsonify({'success': False, 'error': 'No active scrape'}), 400

    scraping_state['paused'] = False
    return jsonify({'success': True, 'message': 'Scraping resumed'}), 200

@app.route('/api/scrape/stop', methods=['POST'])
def stop_scraping():
    """Stop scraping"""
    global scraping_state

    scraping_state['stop_requested'] = True
    scraping_state['active'] = False
    scraping_state['paused'] = False

    return jsonify({'success': True, 'message': 'Scraping stopped'}), 200

@app.route('/api/scrape/status', methods=['GET'])
def get_scrape_status():
    """Get current scraping status"""
    return jsonify({
        'success': True,
        'status': scraping_state
    }), 200

# === LEAD MANAGEMENT ===

@app.route('/api/dashboard/<profile_id>', methods=['GET'])
def get_dashboard_stats(profile_id):
    """Get dashboard statistics for a profile"""
    try:
        profile = profile_manager.get_profile(profile_id)
        if not profile:
            return jsonify({'success': False, 'error': 'Profile not found'}), 404

        db = LeadsDatabase(profile.get_database_path())
        stats = db.get_dashboard_stats()

        return jsonify({
            'success': True,
            'stats': stats
        }), 200
    except Exception as e:
        print(f"[Dashboard] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/leads/<profile_id>', methods=['GET'])
def get_leads(profile_id):
    """Get leads for a profile with optional filtering"""
    try:
        profile = profile_manager.get_profile(profile_id)
        if not profile:
            return jsonify({'success': False, 'error': 'Profile not found'}), 404

        # Get filter parameters
        filter_zip = request.args.get('zip')
        filter_category = request.args.get('category')
        filter_status = request.args.get('status')

        db = LeadsDatabase(profile.get_database_path())

        # Get all leads with status
        leads = db.get_all_leads()

        return jsonify({
            'success': True,
            'leads': [
                {
                    'id': lead[0],  # id from database
                    'name': lead[1],
                    'phone': lead[2],
                    'address': lead[3],
                    'website': lead[4],
                    'email': lead[5],
                    'category': lead[6] if len(lead) > 6 else 'N/A',
                    'zipCode': lead[7] if len(lead) > 7 else 'N/A',
                    'status': lead[8] if len(lead) > 8 else 'New',  # status from database
                    'city': lead[9] if len(lead) > 9 else None  # city/location from database
                }
                for lead in leads
            ]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/leads/<profile_id>/<int:lead_id>/status', methods=['PUT'])
def update_lead_status(profile_id, lead_id):
    """Update the status of a lead"""
    try:
        profile = profile_manager.get_profile(profile_id)
        if not profile:
            return jsonify({'success': False, 'error': 'Profile not found'}), 404

        data = request.json
        new_status = data.get('status')

        if new_status not in ['New', 'Contacted', 'Archived']:
            return jsonify({'success': False, 'error': 'Invalid status'}), 400

        db = LeadsDatabase(profile.get_database_path())
        db.update_lead_status(lead_id, new_status)

        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/leads/<profile_id>/bulk-status', methods=['PUT'])
def bulk_update_lead_status(profile_id):
    """Update status for multiple leads at once"""
    try:
        profile = profile_manager.get_profile(profile_id)
        if not profile:
            return jsonify({'success': False, 'error': 'Profile not found'}), 404

        data = request.json
        lead_ids = data.get('leadIds', [])
        new_status = data.get('status')

        if not lead_ids:
            return jsonify({'success': False, 'error': 'No lead IDs provided'}), 400

        if new_status not in ['New', 'Contacted', 'Archived']:
            return jsonify({'success': False, 'error': 'Invalid status'}), 400

        db = LeadsDatabase(profile.get_database_path())
        updated_count = 0

        for lead_id in lead_ids:
            try:
                db.update_lead_status(lead_id, new_status)
                updated_count += 1
            except Exception as e:
                print(f"[Bulk Update] Failed to update lead {lead_id}: {e}")

        return jsonify({
            'success': True,
            'updated': updated_count,
            'requested': len(lead_ids)
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/leads/<profile_id>/scraped-combos', methods=['GET'])
def get_scraped_combos(profile_id):
    """Get list of ZIP+Category combinations already scraped for this profile"""
    try:
        profile = profile_manager.get_profile(profile_id)
        if not profile:
            return jsonify({'success': False, 'error': 'Profile not found'}), 404

        db = LeadsDatabase(profile.get_database_path())

        # Query unique ZIP + Category combinations from leads
        import sqlite3
        conn = sqlite3.connect(profile.get_database_path())
        cursor = conn.cursor()

        cursor.execute('''
            SELECT DISTINCT zip_code, category, COUNT(*) as lead_count
            FROM leads
            WHERE zip_code IS NOT NULL AND zip_code != '' AND zip_code != 'N/A'
              AND category IS NOT NULL AND category != '' AND category != 'N/A'
            GROUP BY zip_code, category
            ORDER BY zip_code, category
        ''')

        combos = []
        for row in cursor.fetchall():
            combos.append({
                'zip': row[0],
                'category': row[1],
                'leadCount': row[2]
            })

        conn.close()

        return jsonify({
            'success': True,
            'combos': combos,
            'count': len(combos)
        }), 200

    except Exception as e:
        print(f"[Scraped Combos] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/leads/<profile_id>/export', methods=['POST'])
def export_leads(profile_id):
    """Export leads to CSV or XLSX format and save to file"""
    try:
        profile = profile_manager.get_profile(profile_id)
        if not profile:
            return jsonify({'success': False, 'error': 'Profile not found'}), 404

        data = request.json
        lead_ids = data.get('leadIds')  # None for all leads
        file_format = data.get('format', 'csv')
        filename = data.get('filename', 'leads_export')

        db = LeadsDatabase(profile.get_database_path())
        all_leads = db.get_all_leads()

        # Filter leads if specific IDs requested
        if lead_ids:
            leads_dict = {lead[0]: lead for lead in all_leads}
            filtered_leads = [leads_dict[lid] for lid in lead_ids if lid in leads_dict]
        else:
            filtered_leads = all_leads

        # Prepare data
        headers = ['ID', 'Name', 'Phone', 'Address', 'Website', 'Email', 'Category', 'ZIP Code', 'Status']
        rows = []

        for lead in filtered_leads:
            rows.append([
                str(lead[0]),  # ID
                lead[1],  # Name
                lead[2],  # Phone
                lead[3] or 'N/A',  # Address
                lead[4] or 'N/A',  # Website
                lead[5] or 'N/A',  # Email
                lead[6] if len(lead) > 6 else 'N/A',  # Category
                lead[7] if len(lead) > 7 else 'N/A',  # ZIP
                lead[8] if len(lead) > 8 else 'New'  # Status
            ])

        # Create file based on format
        import os
        import tempfile

        if file_format == 'xlsx':
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"

            # Add headers with styling
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            # Add data rows
            for row in rows:
                ws.append(row)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to temp file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            temp_file.close()

            file_path = temp_file.name

        else:  # CSV
            import csv
            temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', newline='', encoding='utf-8')
            writer = csv.writer(temp_file)
            writer.writerow(headers)
            writer.writerows(rows)
            temp_file.close()

            file_path = temp_file.name

        # Read file and return as base64
        import base64
        with open(file_path, 'rb') as f:
            file_data = base64.b64encode(f.read()).decode('utf-8')

        # Clean up temp file
        os.unlink(file_path)

        return jsonify({
            'success': True,
            'fileData': file_data,
            'filename': f"{filename}.{file_format}",
            'count': len(filtered_leads),
            'format': file_format
        }), 200

    except Exception as e:
        print(f"[Export] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# === BACKGROUND SCRAPING LOGIC ===

def add_log(message, log_type='info'):
    """Add a log message to the scraping state"""
    global scraping_state
    import datetime

    # Clean message - remove ANSI codes and emoji that cause issues
    clean_msg = message
    for char in ['🔍', '✓', '✗', '⚠️', '💾', '🧹', '📊']:
        clean_msg = clean_msg.replace(char, '')

    scraping_state['logs'].append({
        'message': clean_msg.strip(),
        'type': log_type,
        'timestamp': datetime.datetime.now().isoformat()
    })

    # Keep only last 100 logs to prevent memory issues
    if len(scraping_state['logs']) > 100:
        scraping_state['logs'] = scraping_state['logs'][-100:]

def run_scrape_job(profile_id, zip_code, category, max_pages):
    """Run scraping job in background thread with real-time logging"""
    global scraping_state

    # Intercept stdout to capture print statements and track page progress
    import sys
    import io
    import re

    class LogCapture(io.StringIO):
        def write(self, message):
            if message.strip():
                add_log(message, 'info')

                # Track current page from log messages
                page_match = re.search(r'\[PAGE (\d+)\]', message)
                if page_match:
                    scraping_state['current_page'] = int(page_match.group(1))
                    # Update progress based on page
                    page_progress = (scraping_state['current_page'] / max_pages) * 70
                    scraping_state['progress'] = 10 + int(page_progress)

            return super().write(message)

    old_stdout = sys.stdout
    sys.stdout = LogCapture()

    try:
        add_log(f"[START] Scraping {category} in ZIP {zip_code}", 'system')
        add_log(f"[CONFIG] Max pages: {max_pages}", 'system')
        scraping_state['progress'] = 10
        scraping_state['current_page'] = 0

        # Back to YellowPages with FIXED Cloudflare bypass (non-headless + longer waits)
        from src.scraper_free_bypass import scrape_yellowpages_free
        leads = scrape_yellowpages_free(zip_code, category, max_pages)

        scraping_state['total_leads'] = len(leads)
        scraping_state['progress'] = 85
        add_log(f"[SUCCESS] Found {len(leads)} businesses", 'success')

        # Save leads to database
        profile = profile_manager.get_profile(profile_id)
        if profile and leads:
            add_log(f"[INFO] Saving leads to database...", 'info')
            scraping_state['progress'] = 90

            # Look up city name from ZIP code
            import pgeocode
            city_name = None
            try:
                nomi = pgeocode.Nominatim('US')
                zip_data = nomi.query_postal_code(zip_code)
                if zip_data is not None and not zip_data.isna().all():
                    city_name = str(zip_data.place_name) if hasattr(zip_data, 'place_name') else None
                    add_log(f"[INFO] City: {city_name}", 'info')
            except Exception as e:
                add_log(f"[WARNING] Could not lookup city for ZIP {zip_code}: {e}", 'info')

            db = LeadsDatabase(profile.get_database_path())
            saved_count = 0
            for lead in leads:
                success, reason = db.add_lead(
                    name=lead['name'],
                    address=lead['address'],
                    phone=lead['phone_number'],
                    email=lead.get('email'),
                    website=lead.get('website'),
                    category=category,
                    zip_code=zip_code,
                    location=city_name
                )
                if success:
                    saved_count += 1

            scraping_state['progress'] = 95
            add_log(f"[INFO] Saved {saved_count} unique leads to database", 'info')

            profile_manager.update_profile_leads(profile_id, db.get_total_leads())
            add_log(f"[COMPLETE] Total unique businesses: {len(leads)}", 'success')

        scraping_state['progress'] = 100

    except Exception as e:
        add_log(f"[ERROR] {str(e)}", 'error')
        scraping_state['error'] = str(e)
        import traceback
        traceback.print_exc()

    finally:
        sys.stdout = old_stdout
        scraping_state['active'] = False
        scraping_state['paused'] = False

# === SERVER STARTUP ===

if __name__ == '__main__':
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 5050
    print(f"[API Server] Starting on http://localhost:{port}")
    print("[API Server] Press Ctrl+C to stop")

    app.run(
        host='127.0.0.1',  # Localhost only for security
        port=port,
        debug=False,
        threaded=True
    )
